from pathlib import Path
import shapefile
from collections import defaultdict
import folium
import json

# Locate project files and folders
ProjectPath = Path(__file__).parent.parent
DataPath = ProjectPath.joinpath('Data')
OutputPath = ProjectPath.joinpath('Output')

shp_path = DataPath.joinpath("Trails_Redlining_30_9/Trails_Redlining_30_9.shp")
redline_xl = DataPath.joinpath("AMC_30th_Edition_Redlining_v30_9.xlsx")
geojson_output = OutputPath.joinpath("trails_geojson.json")
output_map = OutputPath.joinpath("redline_map_xlsx.html")

assert redline_xl.suffix in ['.xls', '.xlsx'], "Spreadsheet must have a .xls or .xlsx extension"

# Define Classes and Functions
class TrailObject:
    """
    Each object represents a trail in the White Mountain Guide 30th Edition.
    """

    def __init__(self):
        self.geom = None
        self.name = None
        self.wmg_section = None
        self.excel_tab = None
        self.mileage = None
        self.miles_to_do = None
        self.identifier = "{0}, {1}".format(self.excel_tab, self.name)


class TabObject:
    """
    Each object represents a tab in the in the excel spreadsheet for redline progress tracking
    """

    def __init__(self, xl_path, xl_ws):
        """
        :param xl_path: pathlib.Path to the spreadsheet
        :param xl_ws: openpyxl or xlrd worksheet
        """
        self.xl_type = xl_path.suffix
        if self.xl_type == '.xlsx':
            self._init_index = 0
            self.wmg_section = xl_ws.cell(1, 1).value
            self.tab_name = xl_ws.title
        else:
            self._init_index = -1
            self.wmg_section = xl_ws.cell_value(0, 0)
            self.tab_name = xl_ws.name
        self.name_column = self._init_index
        self.name_row = self._init_index
        self.last_row = self._init_index
        self.mileage_column = self._init_index
        self.todo_column = self._init_index

    def update_table_bounds(self, xl_ws):
        print("Updating table bounds for {} tab".format(self.tab_name))
        if self.tab_name != 'Summary':
            if self.xl_type == '.xlsx':
                col_num = self._init_index + 1
                while col_num < xl_ws.max_column + 1:
                    row_num = self._init_index + 1
                    while row_num < xl_ws.max_row + 1:
                        if self.name_column == self._init_index:
                            if xl_ws.cell(row_num, col_num).value == 'Trail Name':
                                self.name_column = col_num
                                self.name_row = row_num
                                col_num = xl_ws.max_column + 1
                        row_num += 1
                    col_num += 1
                if self.name_column > self._init_index:
                    col_num = self.name_column + 1
                    while col_num < xl_ws.max_column + 1:
                        if "Total" in xl_ws.cell(self.name_row, col_num).value:
                            self.mileage_column = col_num
                        if "To Do" in xl_ws.cell(self.name_row, col_num).value:
                            self.todo_column = col_num
                        if self.mileage_column > self._init_index and self.todo_column > self._init_index:
                            col_num = self.name_column + 10
                        col_num += 1
                self.last_row = xl_ws.max_row
            else:
                col_num = self._init_index + 1
                while col_num < xl_ws.ncols:
                    # print("Tab = {}; Current Column = {}".format(self.tab_name, col_num))
                    row_num = self._init_index + 1
                    while row_num < xl_ws.nrows:
                        if self.name_column == self._init_index:
                            # print("    Row number: {}".format(row_num))
                            if xl_ws.cell_value(row_num, col_num) == 'Trail Name':
                                self.name_column = col_num
                                self.name_row = row_num
                                col_num = xl_ws.ncols
                        row_num += 1
                    col_num += 1
                if self.name_column > self._init_index:
                    col_num = self.name_column + 1
                    while col_num < xl_ws.ncols:
                        if "Total" in xl_ws.cell_value(self.name_row, col_num):
                            self.mileage_column = col_num
                        if "To Do" in xl_ws.cell_value(self.name_row, col_num):
                            self.todo_column = col_num
                        if self.mileage_column > self._init_index and self.todo_column > self._init_index:
                            col_num = xl_ws.ncols
                        col_num += 1
                self.last_row = xl_ws.nrows - 1

            assert self.name_column > self._init_index, "Check '{0}' tab, trail name column".format(self.tab_name)
            assert self.name_row > self._init_index, "Check '{0}' tab, trail name row".format(self.tab_name)
            assert self.mileage_column > self.name_column, "Check '{0}' tab, mileage column".format(self.tab_name)
            assert self.todo_column > self.name_column, "Check '{0}' tab, to do column".format(self.tab_name)
            assert self.last_row > self.name_row, "Check '{0}' tab, last row: {1}".format(self.tab_name, self.last_row)

            print("Tab = {0} \n"
                  "Name Column = {1} \n"
                  "Name Row = {2} \n"
                  "Mileage Column = {3} \n"
                  "ToDo Column = {4} \n"
                  "Last Row = {5}".format(self.tab_name, self.name_column,
                                          self.name_row, self.mileage_column, self.todo_column, self.last_row))


# Expected tabs in the sheet
tabs_expected = {'Summary': 'White Mountains Redlining Workbook',
                 'Washington': 'Mt. Washington and the Southern Ridges',
                 'Northerns': 'The Northern Peaks and the Great Gulf',
                 'Franconias': 'The Franconia, Twin, and Willey Ranges',
                 'Carrigain': 'The Carrigain and Moat Regions',
                 'Cannon': 'Cannon and Kinsman',
                 'Moosilauke': 'The Moosilauke Region',
                 'Waterville': 'The Waterville Valley and Squam Lake Regions',
                 'Chocorua': 'Mt. Chocorua and the Eastern Sandwich Range',
                 'Carters': 'The Carter and Baldface Ranges',
                 'Speckled': 'Speckled Mountain Region',
                 'Mahoosuc': 'The Mahoosuc Range Area',
                 'NorthernNH': 'Northern New Hampshire'}
attr_dict = dict()
# trails_dict = defaultdict(list)  # [Tab, Trail Name, Mileage, MilesToDo]
# Access redlining spreadsheet and check for format
if redline_xl.suffix == '.xlsx':
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(redline_xl))
    tabs = wb.sheetnames
    assert len(tabs) == len(tabs_expected.keys()), "{0} tabs found. Expected {1}".format(len(tabs), len(tabs_expected))
    for tab in tabs:
        assert tab in tabs_expected.keys(), "Unexpected tab found: '{0}'".format(tab)
        ws = wb[tab]
        assert ws.cell(1, 1).value == tabs_expected[tab], \
            "'{0}' section is '{1}'. Expected '{2}'".format(tab, ws.cell(1, 1).value, tabs_expected[tab])
        tab_object = TabObject(redline_xl, ws)
        tab_object.update_table_bounds(ws)
        i = tab_object.name_row + 1
        while i <= tab_object.last_row:
            _tr_name = ws.cell(i, tab_object.name_column).value
            _tr_id = "{0}, {1}".format(tab, _tr_name)
            _temp_dict = dict()
            _temp_dict["Trail_Name"] = _tr_name
            _temp_dict["Tab"] = tab
            _temp_dict["Section"] = tabs_expected[tab]
            _temp_dict["Mileage"] = ws.cell(i, tab_object.mileage_column).value
            _temp_dict["Miles ToDo"] = ws.cell(i, tab_object.todo_column).value
            attr_dict[_tr_id] = _temp_dict
            i += 1

else:
    import xlrd

    wb = xlrd.open_workbook(str(redline_xl))
    tabs = wb.sheet_names()
    assert len(tabs) == len(tabs_expected.keys()), "{0} tabs found. Expected {1}".format(len(tabs), len(tabs_expected))
    for tab in tabs:
        assert tab in tabs_expected.keys(), "Unexpected tab found: '{0}'".format(tab)
        ws = wb.sheet_by_name(tab)
        assert ws.cell_value(0, 0) == tabs_expected[tab], \
            "'{0}' section is '{1}'. Expected '{2}'".format(tab, ws.cell_value(0, 0), tabs_expected[tab])
        tab_object = TabObject(redline_xl, ws)
        tab_object.update_table_bounds(ws)
        i = tab_object.name_row + 1
        while i <= tab_object.last_row:
            _tr_name = ws.cell_value(i, tab_object.name_column)
            _tr_id = "{0}, {1}".format(tab, _tr_name)
            _temp_dict = dict()
            _temp_dict["Trail_Name"] = _tr_name
            _temp_dict["Tab"] = tab
            _temp_dict["Section"] = tabs_expected[tab]
            _temp_dict["Mileage"] = ws.cell_value(i, tab_object.mileage_column)
            _temp_dict["Miles ToDo"] = ws.cell_value(i, tab_object.todo_column)
            attr_dict[_tr_id] = _temp_dict
            i += 1

# Read shapefile
trail_geom_dict = dict()
reader = shapefile.Reader(str(shp_path))
to_json = []
for sr in reader.shapeRecords():
    # atr = dict(zip(field_names, sr.record))
    trail_id = "{0}, {1}".format(sr.record['Tab'], sr.record['Trail_Name'])
    to_json.append(dict(type="Feature", geometry=sr.shape.__geo_interface__, properties=attr_dict[trail_id]))

# write the GeoJSON file
geojson = open(str(geojson_output), "w")
geojson.write(json.dumps({"type": "FeatureCollection", "features": to_json}, indent=2) + "\n")
geojson.close()

# Create a Map instance
m = folium.Map(location=[44.1, -71.4], tiles='Stamen Terrain', zoom_start=10, control_scale=True)

trail_color = defaultdict(lambda: str('blue'))
trail_color[0] = 'red'


def trail_style(feature):
    return {'opacity': 1,
            'weight': 1,
            'color': trail_color[feature['properties']['Miles ToDo']]}


def highlight_function(feature):
    return {
        'fillColor': '#ffaf00',
        'color': 'yellow',
        'weight': 3
    }


# reading JSON file
with open(str(geojson_output)) as access_json:
    read_content = json.load(access_json)
feature_access = read_content['features']

layer_geom = folium.FeatureGroup(name='Trails', control=True)

for i in range(len(feature_access)):
    temp_geojson = {"features": [feature_access[i]], "type": "FeatureCollection"}
    temp_geojson_layer = folium.GeoJson(temp_geojson, style_function=trail_style, highlight_function=highlight_function)
    trail_name = feature_access[i]["properties"]["Trail_Name"]
    spreadsheet_tab = feature_access[i]["properties"]["Tab"]
    wmg_section = feature_access[i]["properties"]["Section"]
    miles_tot = feature_access[i]["properties"]["Mileage"]
    miles_todo = feature_access[i]["properties"]["Miles ToDo"]
    popup_text = "<strong> Trail Name: </strong>{0}<br>" \
                 "<strong> WMG30 Section: </strong>{1}<br>" \
                 "<strong> Spreadsheet Tab: </strong>{2}<br>" \
                 "<strong> Mileage Total: </strong>{3}<br>" \
                 "<strong> Mileage To Do: </strong>{4}" \
        .format(trail_name, wmg_section, spreadsheet_tab, miles_tot, miles_todo)
    popup = folium.Popup(popup_text, max_width=600)
    popup.add_to(temp_geojson_layer)
    temp_geojson_layer.add_to(layer_geom)

layer_geom.add_to(m)
folium.LayerControl(autoZIndex=False, collapsed=True).add_to(m)

m.save(str(output_map))
