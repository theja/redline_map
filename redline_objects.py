from xlrd import open_workbook
from openpyxl import load_workbook


def difference(lst1, lst2):
    """Get the values in one list but not in other"""
    return [value for value in lst1 if value not in lst2]


class Peak:
    """Peaks in white mountain national forest"""

    def __init__(self):
        self.peak_name = None
        self.elevation = None
        self.is_4k = None
        self.coords = None


class Spreadsheet:
    """This is a redlining spreadsheet used to keep track of hiking progress"""

    def __init__(self, xl_path, tabs_dict):
        """
        @param xl_path: pathlib.Path type object to the excel spreadsheet
        @param tabs_dict: a dictionary with tab names as keys and white mountain guide section names as values
        """
        assert xl_path.suffix in ['.xls', '.xlsx'], "Input spreadsheet must have a '.xls' or '.xlsx' extension"
        self.xl_type = xl_path.suffix
        if xl_path.suffix == '.xls':
            wb = open_workbook(str(xl_path))
            self.wb = wb
            self.tab_names = wb.sheet_names()
            #  check if spreadsheet tabs are as expected
            assert len(difference(list(tabs_dict.keys()), self.tab_names)) == 0, \
                "'{}' tab not found in the spreadsheet".format(difference(list(tabs_dict.keys()), self.tab_names))
            # Find version of workbook. Must be 29th or 30th
            _version = wb.sheet_by_name('Summary').cell_value(1, 0)
            assert ('29' in _version) or ('30' in _version), "Expected 29th or 30th version in cell A2 of Summary tab."
            if '30' in _version:
                self.wmg_edition = 30
            else:
                self.wmg_edition = 29
        else:
            wb = load_workbook(str(xl_path))
            self.wb = wb
            self.tab_names = wb.sheetnames
            #  check if spreadsheet tabs are as expected
            assert len(difference(list(tabs_dict.keys()), self.tab_names)) == 0, \
                "'{}' tab not found in the spreadsheet".format(difference(list(tabs_dict.keys()), self.tab_names))
            # Find version of workbook. Must be 29th or 30th
            _version = wb['Summary'].cell(2, 1).value
            assert ('29' in _version) or ('30' in _version), "Expected 29th or 30th version in cell A2 of Summary tab."
            if '30' in _version:
                self.wmg_edition = 30
            else:
                self.wmg_edition = 29


class TabObject:
    """Represents a tab in the in the excel spreadsheet for redlining workbook spreadsheet"""

    def __init__(self, spreadsheet, tab):
        """
        @param spreadsheet: parent spreadsheet object
        @param tab: tab name in the spreadsheet
        """
        self.xl_type = spreadsheet.xl_type
        self.wb = spreadsheet.wb
        if self.xl_type == '.xlsx':
            self._init_index = 0
            self.ws = self.wb[tab]
        else:
            self._init_index = -1
            self.ws = self.wb.sheet_by_name(tab)
        self.name_column = self._init_index
        self.header_row = self._init_index
        self.last_row = self._init_index
        self.mileage_column = self._init_index
        self.todo_column = self._init_index
        self.trail_attr_dict = dict()  # keys are trails_IDs and values are  tuples (Mileage, Miles To Do)
        self.update_table_bounds(tab)
        self.read_trail_attributes(tab)

    def update_table_bounds(self, tab):
        # print("Updating table bounds for {} tab".format(tab))
        if tab != 'Summary':
            if self.xl_type == '.xlsx':
                col_num = self._init_index + 1
                while col_num < self.ws.max_column + 1:
                    row_num = self._init_index + 1
                    while row_num < self.ws.max_row + 1:
                        if self.name_column == self._init_index:
                            if self.ws.cell(row_num, col_num).value == 'Trail Name':
                                self.name_column = col_num
                                self.header_row = row_num
                                col_num = self.ws.max_column + 1
                        row_num += 1
                    col_num += 1
                if self.name_column > self._init_index:
                    col_num = self.name_column + 1
                    while col_num < self.ws.max_column + 1:
                        if "Total" in self.ws.cell(self.header_row, col_num).value:
                            self.mileage_column = col_num
                        if "To Do" in self.ws.cell(self.header_row, col_num).value:
                            self.todo_column = col_num
                        if self.mileage_column > self._init_index and self.todo_column > self._init_index:
                            col_num = self.name_column + 10
                        col_num += 1
                self.last_row = self.ws.max_row
            else:
                col_num = self._init_index + 1
                while col_num < self.ws.ncols:
                    row_num = self._init_index + 1
                    while row_num < self.ws.nrows:
                        if self.name_column == self._init_index:
                            if self.ws.cell_value(row_num, col_num) == 'Trail Name':
                                self.name_column = col_num
                                self.header_row = row_num
                                col_num = self.ws.ncols
                        row_num += 1
                    col_num += 1
                if self.name_column > self._init_index:
                    col_num = self.name_column + 1
                    while col_num < self.ws.ncols:
                        if "Total" in self.ws.cell_value(self.header_row, col_num):
                            self.mileage_column = col_num
                        if "To Do" in self.ws.cell_value(self.header_row, col_num):
                            self.todo_column = col_num
                        if self.mileage_column > self._init_index and self.todo_column > self._init_index:
                            col_num = self.ws.ncols
                        col_num += 1
                self.last_row = self.ws.nrows - 1

            assert self.name_column > self._init_index, "'{0}' tab, could not find 'Trail Name' column".format(tab)
            assert self.header_row > self._init_index, "'{0}' tab, could not find header row".format(tab)
            assert self.mileage_column > self.name_column, "'{0}' tab, could not find 'Mileage' column".format(tab)
            assert self.todo_column > self.name_column, "'{0}' tab, could not find 'Miles To Do' column".format(tab)
            assert self.last_row > self.header_row, "'{0}' tab, could not find last row of the table".format(tab)

    def read_trail_attributes(self, tab):
        if self.xl_type == '.xlsx':
            i = self.header_row + 1
            while i <= self.last_row:
                _tr_name = self.ws.cell(i, self.name_column).value
                if _tr_name:
                    _tr_id = "{0}, {1}".format(tab, _tr_name)
                    self.trail_attr_dict[_tr_id] = (self.ws.cell(i, self.mileage_column).value,
                                                    self.ws.cell(i, self.todo_column).value)
                i += 1
        else:
            i = self.header_row + 1
            while i <= self.last_row:
                _tr_name = self.ws.cell_value(i, self.name_column)
                if _tr_name:
                    _tr_id = "{0}, {1}".format(tab, _tr_name)
                    self.trail_attr_dict[_tr_id] = (self.ws.cell_value(i, self.mileage_column),
                                                    self.ws.cell_value(i, self.todo_column))
                i += 1


class Trail:
    """Trails in the white mountain guide for """

    def __init__(self, trail_id):
        self.trail_ID = trail_id
        self.alt_ID = None
        self.geom = None
        self.mileage = None
        self.miles_todo = None
        self.tab = self.trail_ID.split(", ")[0]
        self.trail_name = self.trail_ID.split(", ")[1]